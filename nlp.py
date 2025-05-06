''' Job TitleJob_Title_Simplified/Job Description/Sector'''
import numpy as np
import re
import pandas as pd
import nltk
from nltk.corpus import stopwords
from openpyxl import load_workbook
from nltk.stem import PorterStemmer
import seaborn as sns
filepath=r"D:\da\python.xlsx"
df=pd.read_excel(filepath,sheet_name="jd1")
stop_words = set(stopwords.words('english'))
'''stemmer = PorterStemmer()'''
def cleaning(text):
    cleantxt=re.sub('[%s]'%re.escape(""""!@#$%^&*()€™_+<>?:"{}|;'[]\=-"""),' ',text)
    cleantxt = cleantxt.lower()
    words = cleantxt.split()
    filtered_words = [word for word in words if word not in stop_words]
    return ' '.join(filtered_words)
df['jd'] = df['Job Description'].apply(lambda x:cleaning(str(x)))
A=load_workbook(filepath)
B=A["jd1"]
print(df['Job Description'][0])
for index,r in enumerate(df['jd'],start=2):
    B.cell(row=index,column=5,value=str(r))
A.save(filepath)
skills_keywords = [
    "Python", "R", "SQL", "Java", "C++", "MATLAB", "Excel", "Pandas", "NumPy", "SAS", "SPSS",
    "Tableau", "Power BI", "Matplotlib", "Seaborn", "D3.js", "ms access","lms","MachineLearning","machine learning", "DeepLearning","deep learning",
    "TensorFlow", "Keras", "PyTorch", "Scikit-learn", "MySQL","data management fundamentals","software engineering", "PostgreSQL", "Oracle", "MongoDB","math"
    "Hadoop", "Hive", "Spark", "Statistics", "Probability", "Algebra", "Calculus", "Statistical Analysis",
    "BigQuery", "AWS", "Azure", "Google Cloud Platform","cloud" "Git", "Docker", "Kubernetes", "APIs", "ETL",
    "Data Wrangling", "Communication", "Collaboration", "Problem-solving", "Critical Thinking",
    "Attention to Detail","artificialintelligence","ai","artificial intelligence"
]
skills_keywords = [skill.lower() for skill in skills_keywords]
def extract_skills(text, keywords):
    text = text.lower()
    found_skills = [skill for skill in keywords if re.search(r'\b' + re.escape(skill) + r'\b', text)]
    return list(set(found_skills))
df['Extracted Skills'] = df['jd'].apply(lambda x: extract_skills(str(x), skills_keywords))
print(df['Extracted Skills'])
A=load_workbook(filepath)
B=A["jd1"]
for index,r in enumerate(df['Extracted Skills'],start=2):
    B.cell(row=index,column=6,value=str(r))
A.save(filepath)
skills= [
    "Python", "SQL", "artificial intelligence","ai","google cloud platform","deep learning","spark","statistics","machine learning",  "Power BI",
    "Tableau", "ETL","excel","aws",
]
skills = [skill.lower() for skill in skills]
for skill in skills:
    df[skill] = df['Extracted Skills'].apply(lambda skills: 1 if skill in skills else 0)
A = load_workbook(filepath)
B = A["jd1"]
for index, skill in enumerate(skills, start=7):
    skill_column = skill.lower()
    B.cell(row=1, column=index, value=skill)
    for row_idx, value in enumerate(df[skill_column], start=2):
        B.cell(row=row_idx, column=index, value=value)
A.save(filepath)




