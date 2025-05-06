import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Load the Excel file and the specific sheet
filepath = r"D:\da\glassdooranalysis.xlsx"
df = pd.read_excel(filepath, sheet_name="random")

# Perform random sampling
sample_size = 100
sample = df.sample(n=sample_size, random_state=1)
print(sample)

# Open the Excel file and the 'random' sheet
A = load_workbook(filepath)
B = A["random"]

# Write the sample data to the 5th column starting from the 2nd row
for index, r in enumerate(sample.itertuples(), start=2):
    B.cell(row=index, column=5, value=r.Index)

# Save the updated Excel file
A.save(filepath)
