import pandas as pd
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


from pandas import DataFrame

# Load your Excel file
file_path = r"D:\da\jd.xlsx"
sheet_data  = pd.read_excel(file_path, sheet_name="jd")

# Define the skills keywords to search for

skills_keywords = [
    "Python", "R", "SQL", "Java", "C++", "MATLAB", "Excel", "Pandas", "NumPy", "SAS", "SPSS",
    "Tableau", "Power BI", "Matplotlib", "Seaborn", "D3.js", "ms access","lms","MachineLearning","machine-learning", "DeepLearning",
    "TensorFlow", "Keras", "PyTorch", "Scikit-learn", "MySQL", "PostgreSQL", "Oracle", "MongoDB","math"
    "Hadoop", "Hive", "Spark", "Statistics", "Probability", "Algebra", "Calculus", "Statistical Analysis",
    "BigQuery", "AWS", "Azure", "Google Cloud Platform", "Git", "Docker", "Kubernetes", "APIs", "ETL",
    "Data Wrangling", "Communication","communication skills", "Collaboration", "Problem-solving", "Critical Thinking",
    "Attention to Detail"
]

# Convert keywords to lowercase
skills_keywords = [skill.lower() for skill in skills_keywords]


# Function to extract keywords using regular expressions
def extract_skills(text, keywords):
    text = text.lower()
    found_skills = [skill for skill in keywords if re.search(r'\b' + re.escape(skill) + r'\b', text)]
    return list(set(found_skills))

# Apply the extraction
sheet_data['Extracted Skills'] = sheet_data['Job Description'].apply(lambda x: extract_skills(str(x), skills_keywords))

# Display the results
print(sheet_data.columns)
print(sheet_data['Job Description'])

print(sheet_data[['Job Title', 'Extracted Skills']])
A=load_workbook(file_path)
B=A["jd"]
for index,r in enumerate(sheet_data['Extracted Skills'],start=2):
    B.cell(row=index,column=4,value=str(r))
A.save(file_path)




