import pandas as pd
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
filepath = r"D:\da\glassdooranalysis.xlsx"
df = pd.read_excel(filepath, sheet_name="cluster")
features = df[['Rating', 'average']]
scaler = StandardScaler()
scaled_features = scaler.fit_transform(features)
kmeans = KMeans(n_clusters=6, random_state=0)
df['Cluster'] = kmeans.fit_predict(scaled_features)
print(df)
A=load_workbook(filepath)
B=A["cluster"]
for index,r in enumerate(df['Cluster'],start=2):
    B.cell(row=index,column=9,value=r)
A.save(filepath)
plt.scatter(df['Rating'], df['average'], c=df['Cluster'], cmap='viridis', s=100)
plt.xlabel('Rating')
plt.ylabel('Avg_Salary')
plt.title('Clusters of Companies')
plt.axvline(x=3.7, color='red', linestyle='--', label='Rating Divider')
plt.axhline(y=80, color='red', linestyle='--', label='salary Divider')
plt.axhline(y=145, color='red', linestyle='--', label='salary Divider')
plt.axhline(y=220, color='red', linestyle='--', label='salary Divider')
plt.legend()
plt.show()
